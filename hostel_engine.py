"""
hostel_engine.py
================
Core business logic for the Hostel Room Allocation System.

All data lives in hostel_data.xlsx (three sheets: Students, Rooms, Allocation).
This module reads/writes that file and enforces every business rule:

  * Double-occupancy rooms (max 2 students per room)
  * Gender must match room's Gender Allowed
  * AC Preference must match room's Room Type
  * First-Come-First-Serve (lowest room number first)
  * No manual row deletion needed — vacate() handles removals cleanly

Usage (standalone, no UI):
    from hostel_engine import HostelSystem
    hs = HostelSystem("hostel_data.xlsx")
    result = hs.allocate("GIM011", "Vikram Bose", "Male", "AC")
    print(result)          # {"status": "success", "room": "101", ...}
    hs.vacate(["GIM001"])
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styling helpers (keeps Excel output clean) ────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
BODY_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def _style_ws(ws, col_widths):
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for r_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            cell.border    = THIN
            cell.alignment = CENTER
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = BODY_FONT
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Main engine ───────────────────────────────────────────────────────────────

class HostelSystem:
    """
    Manages hostel room allocations backed by an Excel file.

    Parameters
    ----------
    excel_path : str | Path
        Path to hostel_data.xlsx.  Must have sheets: Students, Rooms, Allocation.
    """

    MAX_OCCUPANCY = 2  # double-occupancy rule

    def __init__(self, excel_path: str):
        self.path = Path(excel_path)
        if not self.path.exists():
            raise FileNotFoundError(f"Data file not found: {self.path}")

    # ── internal data loaders ─────────────────────────────────────────────────

    def _load_rooms(self) -> pd.DataFrame:
        """Return rooms DataFrame with normalised text columns."""
        df = pd.read_excel(self.path, sheet_name="Rooms", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        df["Room No"]        = df["Room No"].str.strip()
        df["Room Type"]      = df["Room Type"].str.strip().str.title()    # AC / Non-Ac → normalise
        df["Gender Allowed"] = df["Gender Allowed"].str.strip().str.title()
        # normalise "Non-AC" / "Non-ac" / "non-ac" → "Non-AC"
        df["Room Type"] = df["Room Type"].apply(_normalise_type)
        return df

    def _load_allocation(self) -> pd.DataFrame:
        """Return current allocation DataFrame (may be empty)."""
        df = pd.read_excel(self.path, sheet_name="Allocation", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        df = df.dropna(subset=["GIM ID"])
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()
        return df

    def _load_students(self) -> pd.DataFrame:
        df = pd.read_excel(self.path, sheet_name="Students", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        df = df.dropna(subset=["GIM ID"])
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()
        df["Gender"]        = df["Gender"].str.title()
        df["AC Preference"] = df["AC Preference"].apply(_normalise_type)
        return df

    # ── public API ────────────────────────────────────────────────────────────

    def allocate(
        self,
        gim_id: str,
        name: str,
        gender: str,
        ac_preference: str,
    ) -> dict:
        """
        Allocate a room to a new student.

        Returns a dict with keys:
            status  : "success" | "already_allocated" | "no_room" | "error"
            message : human-readable description
            room    : room number (only on success)
        """
        gim_id       = gim_id.strip().upper()
        name         = name.strip()
        gender       = gender.strip().title()
        ac_preference = _normalise_type(ac_preference.strip())

        # ── input validation ──────────────────────────────────────────────────
        if gender not in ("Male", "Female"):
            return _err("Gender must be 'Male' or 'Female'.")
        if ac_preference not in ("AC", "Non-AC"):
            return _err("AC Preference must be 'AC' or 'Non-AC'.")

        rooms      = self._load_rooms()
        allocation = self._load_allocation()

        # ── already allocated? ────────────────────────────────────────────────
        if gim_id in allocation["GIM ID"].values:
            existing_room = allocation.loc[allocation["GIM ID"] == gim_id, "Room No"].iloc[0]
            return {
                "status":  "already_allocated",
                "message": f"{gim_id} is already allocated to Room {existing_room}.",
                "room":    existing_room,
            }

        # ── occupancy count per room ──────────────────────────────────────────
        occupancy = (
            allocation.groupby("Room No")
            .size()
            .reset_index(name="occupants")
        )

        # ── candidate rooms ───────────────────────────────────────────────────
        candidates = rooms[
            (rooms["Gender Allowed"] == gender) &
            (rooms["Room Type"]      == ac_preference)
        ].copy()

        if candidates.empty:
            return {
                "status":  "no_room",
                "message": f"No {ac_preference} {gender} rooms exist in the hostel.",
            }

        # merge with occupancy
        candidates = candidates.merge(occupancy, on="Room No", how="left")
        candidates["occupants"] = candidates["occupants"].fillna(0).astype(int)

        # filter rooms with space
        available = candidates[candidates["occupants"] < self.MAX_OCCUPANCY]

        if available.empty:
            return {
                "status":  "no_room",
                "message": (
                    f"No {ac_preference} {gender} room is available right now. "
                    "All matching rooms are full."
                ),
            }

        # ── pick lowest room number ───────────────────────────────────────────
        # sort as numbers if possible, else lexicographically
        try:
            available = available.assign(_sort_key=available["Room No"].astype(int))
        except ValueError:
            available = available.assign(_sort_key=available["Room No"])
        available = available.sort_values("_sort_key")
        chosen_room = available.iloc[0]["Room No"]
        room_type   = available.iloc[0]["Room Type"]

        # ── write back ────────────────────────────────────────────────────────
        new_row = {
            "GIM ID":        gim_id,
            "Student Name":  name,
            "Gender":        gender,
            "AC Preference": ac_preference,
            "Room No":       chosen_room,
            "Room Type":     room_type,
            "Allocated On":  datetime.now().strftime("%d-%b-%Y %H:%M"),
        }
        allocation = pd.concat([allocation, pd.DataFrame([new_row])], ignore_index=True)
        self._save_allocation(allocation)

        return {
            "status":  "success",
            "message": f"✅  Room {chosen_room} ({room_type}) allocated to {name} ({gim_id}).",
            "room":    chosen_room,
        }

    def vacate(self, gim_ids: list) -> dict:
        """
        Remove allocations for a list of GIM IDs (students who have left).

        Returns a dict with keys:
            vacated    : list of GIM IDs successfully vacated
            not_found  : list of GIM IDs that had no allocation
            message    : summary string
        """
        gim_ids    = [g.strip().upper() for g in gim_ids if g.strip()]
        allocation = self._load_allocation()

        existing    = set(allocation["GIM ID"].values)
        vacated     = [g for g in gim_ids if g in existing]
        not_found   = [g for g in gim_ids if g not in existing]

        if vacated:
            allocation = allocation[~allocation["GIM ID"].isin(vacated)].reset_index(drop=True)
            self._save_allocation(allocation)

        parts = []
        if vacated:
            parts.append(f"✅  Vacated: {', '.join(vacated)}")
        if not_found:
            parts.append(f"⚠️  Not found in allocation: {', '.join(not_found)}")

        return {
            "vacated":   vacated,
            "not_found": not_found,
            "message":   "  |  ".join(parts) if parts else "Nothing to do.",
        }

    def get_vacancy_summary(self) -> pd.DataFrame:
        """
        Return a summary DataFrame of available beds by Room Type and Gender.

        Columns: Room Type, Gender, Total Rooms, Total Beds, Occupied, Available
        """
        rooms      = self._load_rooms()
        allocation = self._load_allocation()

        occupancy = allocation.groupby("Room No").size().reset_index(name="occupants")
        rooms     = rooms.merge(occupancy, on="Room No", how="left")
        rooms["occupants"] = rooms["occupants"].fillna(0).astype(int)

        summary = (
            rooms.groupby(["Room Type", "Gender Allowed"])
            .agg(
                Total_Rooms=("Room No", "count"),
                Occupied_Beds=("occupants", "sum"),
            )
            .reset_index()
        )
        summary.rename(columns={"Gender Allowed": "Gender"}, inplace=True)
        summary["Total_Beds"]      = summary["Total_Rooms"] * self.MAX_OCCUPANCY
        summary["Available_Beds"]  = summary["Total_Beds"] - summary["Occupied_Beds"]
        return summary[["Room Type", "Gender", "Total_Rooms", "Total_Beds", "Occupied_Beds", "Available_Beds"]]

    def lookup_student(self, gim_id: str):
        """
        Look up a student in the Students sheet by GIM ID.
        Returns a dict with Name, Gender, AC Preference — or None if not found.
        """
        gim_id   = gim_id.strip().upper()
        students = self._load_students()
        match    = students[students["GIM ID"].str.upper() == gim_id]
        if match.empty:
            return None
        row = match.iloc[0]
        return {
            "Name":          row["Name"],
            "Gender":        row["Gender"],
            "AC Preference": row["AC Preference"],
        }

    def get_current_allocation(self) -> pd.DataFrame:
        """Return the full current allocation as a DataFrame."""
        return self._load_allocation()

    def get_room_detail(self) -> pd.DataFrame:
        """Return per-room occupancy detail."""
        rooms      = self._load_rooms()
        allocation = self._load_allocation()
        occupancy  = (
            allocation.groupby("Room No")["GIM ID"]
            .apply(lambda x: ", ".join(x))
            .reset_index(name="Students")
        )
        occupancy2 = allocation.groupby("Room No").size().reset_index(name="Occupants")
        rooms = rooms.merge(occupancy2, on="Room No", how="left")
        rooms = rooms.merge(occupancy,  on="Room No", how="left")
        rooms["Occupants"] = rooms["Occupants"].fillna(0).astype(int)
        rooms["Students"]  = rooms["Students"].fillna("—")
        rooms["Status"]    = rooms["Occupants"].apply(
            lambda x: "Full" if x >= self.MAX_OCCUPANCY else ("Partial" if x > 0 else "Vacant")
        )
        return rooms

    # ── internal writer ───────────────────────────────────────────────────────

    def _save_allocation(self, df: pd.DataFrame):
        """Re-write only the Allocation sheet; preserve other sheets."""
        wb = openpyxl.load_workbook(self.path)

        # remove old sheet
        if "Allocation" in wb.sheetnames:
            del wb["Allocation"]

        ws = wb.create_sheet("Allocation")

        headers = ["GIM ID", "Student Name", "Gender", "AC Preference",
                   "Room No", "Room Type", "Allocated On"]
        ws.append(headers)

        for _, row in df.iterrows():
            ws.append([row.get(h, "") for h in headers])

        _style_ws(ws, [14, 22, 12, 16, 12, 14, 20])

        # keep sheet order: Students, Rooms, Allocation
        desired = ["Students", "Rooms", "Allocation"]
        current = wb.sheetnames
        ordered = [s for s in desired if s in current] + [s for s in current if s not in desired]
        wb._sheets = [wb[s] for s in ordered]

        wb.save(self.path)


# ── small helpers ─────────────────────────────────────────────────────────────

def _normalise_type(value: str) -> str:
    """Normalise 'non-ac', 'Non-Ac', 'non ac' → 'Non-AC'; 'ac' → 'AC'."""
    v = value.strip().upper().replace(" ", "-")
    if v in ("AC", "A/C"):
        return "AC"
    if v in ("NON-AC", "NON-A/C", "NONAC"):
        return "Non-AC"
    # return title-cased original if unknown (will fail validation later)
    return value.strip().title()

def _err(msg: str) -> dict:
    return {"status": "error", "message": f"❌  {msg}"}
